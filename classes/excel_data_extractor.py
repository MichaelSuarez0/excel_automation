import re
import os
import pandas as pd
import openpyxl
from icecream import ic
import logging
import datetime
import pandas as pd
from typing import Tuple, Optional


script_dir = os.path.abspath(os.path.dirname(__file__))
#macros_folder = os.path.join(script_dir, "..", "macros", "excel")
save_dir = os.path.join(script_dir, "..", "charts")

class ExcelDataExtractor():
    def __init__(self, file_name: str):
        """Class to obtain data from an Excel file, convert to DataFrame, apply transformations, and export it. 
        Engine: mostly pandas

        Parameters
        ----------
        file_name : str
                The name of the Excel file to be loaded (from databases folder)
        """
        self.file_path = os.path.join(script_dir, "..", "databases", f'{file_name}.xlsx')
        self.output_path = os.path.join(script_dir, "..", "charts")
        self.wb = None
        self.ws = None
        self.load_workbook()
    
    def load_workbook(self):
        self.wb = openpyxl.load_workbook(self.file_path)
        # Access different worksheets with self.wb.sheetnames[int]
        self.ws= self.wb.active
    
    def save_workbook(self, output_name: str)-> None:
        """Save your workbook. Automatically includes extension in the name if not declared.

        Args:
            name (str, optional): Choose a name for your Excel file. Defaults to "excel_test".
        """
        self.wb.save(self.output_path)
        print(f'✅ Excel guardado como "{self.output_path}"')

    # def open_new_workbook(self, ws_name: str = None) -> Tuple[Workbook, Worksheet]:
    #     """Dynamically create new workbooks and name them wb2, wb3, etc."""        
    #     self.wb_count += 1  
        
    #     # Create new workbook and assign it dynamically
    #     new_wb_name = f"wb{self.wb_count}"
    #     self.workbooks[(self.wb_count)] = Workbook()
        
    #     # Create new variables dynamically (starting with .self)
    #     setattr(self, new_wb_name, self.workbooks[self.wb_count])
    #     setattr(self, f"ws{self.wb_count}", self.workbooks[self.wb_count].active)
    #     # Get the active worksheet or create a new one with the specified name
    #     if ws_name:
    #         new_ws = self.workbooks[self.wb_count].create_sheet(title=ws_name)
    #     else:
    #         new_ws = self.workbooks[self.wb_count].active

    #     print(f"✅ Created new workbook: {new_wb_name}")
    #     return self.workbooks[self.wb_count], new_ws
    
    @property
    def sheet_names(self) -> list: 
        """Devuelve una lista de los nombres de las hojas."""
        print("Sheet names:")
        for sheet_name in self.wb.sheetnames:
            print(f"- {sheet_name}")
        return self.wb.sheetnames

    @property
    def count_sheets(self) -> int:
        count = len(self.wb.sheetnames)
        print(f'The workbook has {count} sheets.')
        return count
    
    # Reading methods
    def worksheet_to_dataframe(self, sheet_index: int = None) -> pd.DataFrame:
        """Reads sheet and return a DataFrame, may specify worksheet index"""
        sheet_name = self.wb.sheetnames[sheet_index] if sheet_index else self.wb.sheetnames[0]
        df = pd.read_excel(self.file_path, sheet_name)
        return df
    
    def worksheets_to_dataframes(self, include_first = False) -> list[pd.DataFrame]:
        """Reads all sheets at once and returns a list of DataFrames, may specify to skip first"""
        dfs_dict = pd.read_excel(self.file_path, sheet_name=None) # This method reads all sheets at once a returns a dictionary of DataFrames
        sheet_names = list(dfs_dict.keys())[1:] if not include_first else list(dfs_dict.keys())
        dfs = [dfs_dict[name] for name in sheet_names]
        return dfs
    
    # Transformation methods
    def normalize_orientation(self, dfs: pd.DataFrame | list[pd.DataFrame]) -> list[pd.DataFrame]:
        """
        Normalizes the orientation of one or more DataFrames. Converts to list if a single DataFrame is provided.

        Parameters:
        ----------
        dfs : pd.DataFrame | list[pd.DataFrame]
            A DataFrame or a list of DataFrames to be normalized.

        Returns:
        -------
        list[pd.DataFrame]
            A list of normalized DataFrames with the correct orientation.
        
        Raises:
        ------
        ValueError
            If the input is neither a DataFrame nor a list of DataFrames.
        """
        if not isinstance(dfs, (pd.DataFrame, list)):
            raise ValueError("Must provide either a DataFrame or a list of DataFrames")
        if isinstance(dfs, pd.DataFrame):
            dfs= [dfs]
        normalized_dfs = []
        for df in dfs:
            # Check if the first row contains categories. If it does, it will transpose the df.
            if not isinstance(df.iloc[0, 1], str):
                index_name = df.columns[0]
                df = df.set_index(df.columns[0]).transpose() # Manually set another index, or else the default index stays on top
                df.reset_index(inplace=True)
                df.columns = [index_name] + df.columns[1:].tolist()  # I loathe pandas indexes
            df[df.columns[0]] = df[df.columns[0]].astype(str).apply(lambda x: x.strip()) # Clean blank spaces
            normalized_dfs.append(df)
        
        return normalized_dfs
    
    def filter_data(
        self,
        df: pd.DataFrame,
        selected_categories: Optional[list[str]] = None,
        filter_out: Optional[list[str]] = None
    ) -> pd.DataFrame:
        """
        Filters the input DataFrame by selecting columns based on the provided categories.
        Should be run AFTER normalizing orientation for all DataFrames.

        This function filters the given DataFrame (`df`) by selecting only columns (containing
        categories like Departamento) that match the values in `selected_categories`. It
        ensures no duplicated columns are selected, and always includes the first column of 
        the DataFrame (typically used as an identifier or key). If no `selected_categories` 
        are provided, the function returns the original DataFrame without any filtering.

        Parameters:
        ----------
        df : pd.DataFrame
            The DataFrame to be filtered.
        selected_categories : list[str], optional
            A list of column names to be included in the filtered DataFrame. If `None`, 
            no filtering is applied, and the original DataFrame is returned.

        Returns:
        -------
        pd.DataFrame
            A DataFrame containing only the selected columns, including the first column.
        
        Example:
        --------
        # Sample usage:
        df = pd.DataFrame({
            'Departamento': ['Lima', 'Arequipa', 'Cusco'],
            '2014': [10, 20, 30],
            '2015': [11, 21, 31]
        })

        selected_categories = ['Lima']
        filtered_df = filter_data(df, selected_categories)
        print(filtered_df)

        Output:
        --------
        Departamento   2014 2015
        0   Lima        10  11

        """
        if selected_categories:
            cols = [df.columns[0]] # Labels are in Row 1
            
            # Loop through the remaining columns and add them if they are in selected_labels
            for col in df.columns[1:]:
                if col in selected_categories and col not in cols:  # Asegúrate de no añadir duplicados
                    cols.append(col)
            filtered_df = df[cols]
            
        if filter_out:
            cols = [df.columns[0]] # Labels are in Row 1
            
            # Loop through the remaining columns and add them if they are in selected_labels
            for col in df.columns[1:]:
                if col not in filter_out and col not in cols:  # Asegúrate de no añadir duplicados
                    cols.append(col)
            filtered_df = df[cols]

        return filtered_df

    
    def concat_dataframes(
        self, 
        df1: pd.DataFrame,
        df2: pd.DataFrame,
        df1_name: str,
        df2_name: str,
    )-> pd.DataFrame:
        """Concatenates two DataFrames and adds a row for the 'Tipo' at the start of the DataFrame."""

        df1 = pd.concat([df1, pd.DataFrame([['Tipo'] + [df1_name] * (len(df1.columns)-1)], columns=df1.columns)]).reset_index(drop=True)
        df2 = pd.concat([df2, pd.DataFrame([['Tipo'] + [df2_name] * (len(df2.columns)-1)], columns=df2.columns)]).reset_index(drop=True)
        first_col_df1 = df1.columns[0]
        first_col_df2 = df2.columns[0]

        # Check if the first columns match before merging
        if first_col_df1 != first_col_df2:
            raise KeyError(f"The first columns do not match: '{first_col_df1}' and '{first_col_df2}'")
        result_df = pd.merge(df1, df2, on=first_col_df1, how='outer', suffixes=('_1', '_2'))

        # Mover la fila "Tipo" al principio
        result_df = pd.concat([result_df.iloc[[-1],:], result_df.drop(result_df.index[-1], axis=0)], axis=0)

        # Convertir la primera fila (Tipo) en los nombres de las columnas
        tipo_row = result_df[result_df.iloc[:, 0] == 'Tipo'].iloc[0]  # Encontrar la fila 'Tipo'
        result_df.columns = tipo_row 

        # Eliminar la fila "Tipo"
        result_df = result_df[result_df.iloc[:, 0] != 'Tipo'].reset_index(drop=True)

        return result_df

    # Writing methods (simple)
    def dataframe_to_worksheet(self, df: pd.DataFrame, output_name: str, sheet_name: str = 'Hoja1', mode: str = 'w') -> None:
        """Writes a DataFrame to a worksheet in the Excel file.

        Parameters
        ----------
        df : pd.DataFrame
            The DataFrame to write to the worksheet.
        sheet_name : str, optional
            The name of the worksheet. Defaults to 'Hoja1'.
        mode : str, optional
            The mode to open the Excel file ('w' for write, 'a' for append). Defaults to 'w'.
        """
        output_file_path = os.path.join(self.output_path, f'{output_name}.xlsx')
        with pd.ExcelWriter(output_file_path, engine='openpyxl', mode=mode) as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        
    def dataframes_to_worksheets(
        self,
        dfs: list[pd.DataFrame],
        output_name: str,
        sheet_names: list[str] = None, 
        mode: str = 'w', 
        skip_first: bool = True
    ) -> None:
        """Writes multiple DataFrames to multiple worksheets in the Excel file.

        Parameters
        ----------
        dfs : List[pd.DataFrame]
            A list of DataFrames to write to the worksheets.
        sheet_names : list[str], optional
            A list of worksheet names. If not provided, default names will be used.
        mode : str, optional
            The mode to open the Excel file ('w' for write, 'a' for append). Defaults to 'w'.
        skip_first : bool, optional
            Whether to start writing from Worksheet 2 onward. Defaults to True.
        """
        if sheet_names is None:
            sheet_names = [f'Hoja{i+1}' for i in range(len(dfs))]  # Default sheet names: Hoja1, Hoja2, etc.

        if len(dfs) != len(sheet_names):
            raise ValueError("The number of DataFrames must match the number of sheet names.")

        output_file_path = os.path.join(self.output_path, f'{output_name}.xlsx')

        with pd.ExcelWriter(output_file_path, engine='openpyxl', mode=mode) as writer:
            # If skip_first is True, create an empty sheet as the first one
            if skip_first:
                pd.DataFrame().to_excel(writer, sheet_name='Índice')

            # Write DataFrames to subsequent sheets
            for df, sheet_name in zip(dfs, sheet_names):
                df.to_excel(writer, sheet_name=sheet_name, index=False)
   
    # TODO: All that is missing is FUENTE and URL
    # TODO: Use sheet index instead of name



