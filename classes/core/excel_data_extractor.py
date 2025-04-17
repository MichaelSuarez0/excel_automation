import os
import pandas as pd
import openpyxl
import pandas as pd
from typing import Literal, Optional

script_dir = os.path.abspath(os.path.dirname(__file__))

class ExcelDataExtractor():
    def __init__(self, file_name: str, folder: str = "otros", custom_path: str = ""):
        """Class to obtain data from an Excel file, convert to DataFrame, apply transformations, and export it. 
        Engine: mostly pandas

        Parameters
        ----------
        file_name : str
                The name of the Excel file to be loaded (from databases folder)
        folder : str, optional:
            Folder name inside "databases" where file is located (defaults to "otros")
        custom_path : str, optional
            If provided, this path (starting from base dir) is used instead of constructing a path based on 'databases' + 'folder'.
        """
        if custom_path:
            self.file_path = os.path.join(custom_path,  f'{file_name}.xlsx')
        else:
            self.file_path = os.path.join(script_dir, "..", "..", "databases", folder, f'{file_name}.xlsx')
        self.output_path = os.path.join(script_dir, "..", "..", "products")
        self.wb = None
        self.ws = None
        self.load_workbook()
    
    def load_workbook(self):
        self.wb = openpyxl.load_workbook(self.file_path)
        # Access different worksheets with self.wb.sheetnames[int]
        self.ws= self.wb.active
    
    @property
    def sheet_names(self) -> list: 
        """Devuelve una lista de los nombres de las hojas."""
        print("Sheet names:")
        for sheet_name in self.wb.sheetnames:
            print(f"- {sheet_name}")
        return self.wb.sheetnames

    @property
    def count_sheets(self) -> int:
        count = len(self.wb.sheetnames)
        print(f'The workbook has {count} sheets.')
        return count
    
    # Reading methods
    def _preprocess_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        # Remove rows/columns that are completely empty, then replace NaN with ""
        df = df.dropna(axis=0, thresh=1).dropna(axis=1, thresh=1).fillna("")
        
        # Remove extra spaces from string columns
        object_columns = df.select_dtypes(include=['object']).columns
        for col in object_columns:
            df[col] = df[col].apply(lambda x: x.strip() if isinstance(x, str) else x)
        
        return df

    def worksheet_to_dataframe(self, sheet_index: int = None) -> pd.DataFrame:
        """
        Reads a single worksheet and returns it as a cleaned DataFrame.
        
        Parameters
        ----------
        sheet_index : int, optional
            The index of the worksheet to read. If not provided, the first sheet is used.
        """
        sheet_name = self.wb.sheetnames[sheet_index] if sheet_index is not None else self.wb.sheetnames[0]
        df = pd.read_excel(self.file_path, sheet_name=sheet_name)
        return self._preprocess_dataframe(df)

    def worksheets_to_dataframes(self, include_first: bool = True) -> list[pd.DataFrame]:
        """
        Reads all worksheets at once and returns a list of cleaned DataFrames.
        
        Parameters
        ----------
        include_first : bool, optional
            Whether to include the first worksheet. By default, the first worksheet is skipped.
        """
        dfs_dict = pd.read_excel(self.file_path, sheet_name=None) # This pandas funct returns a dictionary: {sheet_name: DataFrame}
        sheet_names = list(dfs_dict.keys())[1:] if not include_first else list(dfs_dict.keys()) # Select the sheet names based on whether the first sheet should be included
        return [self._preprocess_dataframe(dfs_dict[name]) for name in sheet_names]
    
    # Transformation methods
    def normalize_orientation(self, dfs: pd.DataFrame | list[pd.DataFrame]) -> pd.DataFrame | list[pd.DataFrame]:
        """
        Normalizes the orientation of one or more DataFrames. Converts to list if a single DataFrame is provided.

        Parameters:
        ----------
        dfs : pd.DataFrame | list[pd.DataFrame]
            A DataFrame or a list of DataFrames to be normalized.

        Returns:
        -------
        pd.DataFrame | list[pd.DataFrame]
            A normalized DataFrame if a single DataFrame was passed, or a list of normalized DataFrames if a list was passed.
        
        Raises:
        ------
        ValueError
            If the input is neither a DataFrame nor a list of DataFrames.
        """
        if not isinstance(dfs, (pd.DataFrame, list)):
            raise ValueError("Must provide either a DataFrame or a list of DataFrames")
        if isinstance(dfs, pd.DataFrame):
            dfs= [dfs]
        normalized_dfs = []
        for df in dfs:
            # If headers are not strings (categories), df will be transposed.
            #if isinstance(df.iloc[0,1], str) and not df.shape[1] < 3: # Consider using as well if not isinstance(df.columns[1], str)
            index_name = df.columns[0]
            df = df.set_index(df.columns[0]).transpose() # Manually set another index, or else the default index stays on top
            df.reset_index(inplace=True)
            df.columns = [index_name] + df.columns[1:].tolist()  # I loathe pandas indexes
            normalized_dfs.append(df)
        
        # If a single DataFrame was passed, return just that DataFrame.
        if len(normalized_dfs) == 1:
            return normalized_dfs[0]
        
        return normalized_dfs
    
    # TODO: Raise KeyError if at least 1 selected category is not found
    def filter_data(
        self,
        df: pd.DataFrame | list[pd.DataFrame],
        selected_categories: list[str] | str,
        filter_out: bool = False,
        key: Literal["row", "column"] = "column"
    ) -> pd.DataFrame | list[pd.DataFrame]:
        """
        Filters DataFrame(s) by selecting or excluding specific columns or rows.

        This function filters `df` by keeping columns or rows in `selected_categories`. 
        Alternatively, it can also filter out `selected_categories`.

        Parameters
        ----------
        df : pd.DataFrame or list[pd.DataFrame]
            The DataFrame(s) to filter.
        selected_categories : list[str], optional
            A list of column names or row values to include/exclude. If `None`, no filtering is applied.
        filter_out : bool, default=False
            If `True`, excludes the specified columns/rows instead of including them.
        key : {"row", "column"}, default="column"
            Determines whether filtering is applied to columns or rows (based on the first column).

        Returns
        -------
        pd.DataFrame or list[pd.DataFrame]
            Filtered DataFrame(s). Returns the same type as input (single df or list).

        Raises
        ------
        KeyError
            If at least one column or row is not found in the df.
        ValueError
            If key is not "row" or "column".

        Examples
        --------
        >>> df = pd.DataFrame({
        ...     "Departamento": ["Lima", "Arequipa", "Cusco"],
        ...     "2014": [10, 20, 30],
        ...     "2015": [11, 21, 31]
        ... })
        
        **Filtering columns:**
        >>> filter_data(df, ['2014'], key="column")
        Departamento  2014
        0         Lima    10
        1     Arequipa    20
        2        Cusco    30
        
        **Filtering rows:**
        >>> filter_data(df, ['Lima'], key="row")
        Departamento  2014  2015
        0         Lima    10    11
        """
        is_single_df = isinstance(df, pd.DataFrame)
        dfs = [df] if is_single_df else df.copy()
        
        # Normalize selected_categories to list
        if isinstance(selected_categories, str):
            selected_categories = [selected_categories]
        
        filtered_dfs = []
        for df_item in dfs:
            if key == "column":
                missing_categories = [cat for cat in selected_categories if cat not in df_item.columns[1:]]
                if missing_categories:
                    raise KeyError(f"Some columns do not exist in the DataFrame, check typing: {missing_categories}")
                
                if not filter_out:
                    cols = [df_item.columns[0]] + [
                        col for col in selected_categories 
                        if col in df_item.columns and col != df_item.columns[0]]
                    result = df_item[cols]
                else:
                    # Filtrar excluyendo
                    cols = [df_item.columns[0]] + [
                        col for col in df_item.columns[1:] 
                        if col not in selected_categories
                    ]
                    result = df_item[cols]
                if len(cols) == 1:  # Only first column remains
                    raise KeyError(f"No columns in {selected_categories} matched.")
                    
            elif key == "row":
                # Filtrar y mantener el orden exacto de selected_categories
                # ordered_categories = [cat for cat in selected_categories 
                #                     if cat in df_item.iloc[:, 0].values]
                
                # if not ordered_categories and not filter_out:
                #     raise KeyError(f"No rows matched: {selected_categories}")
                missing_categories = [cat for cat in selected_categories if cat not in df_item.iloc[:, 0].values]
                if missing_categories:
                    raise KeyError(f"Some rows do not exist in the DataFrame, check typing: {missing_categories}")

                # Crear máscara y ordenar
                if not filter_out:
                    result = df_item[df_item.iloc[:, 0].isin(selected_categories)]
                    result = result.set_index(result.columns[0]).loc[selected_categories].reset_index()    
                else:
                    result = df_item[~df_item.iloc[:, 0].isin(selected_categories)]
                
                if result.empty:
                    raise KeyError(f"No rows matched: {selected_categories}")
            
            filtered_dfs.append(result)
        
        # Return single df if input was single, else return list
        return filtered_dfs[0] if is_single_df else filtered_dfs

    
    def concat_dataframes(
        self, 
        df1: pd.DataFrame,
        df2: pd.DataFrame,
        df1_name: str,
        df2_name: str,
    )-> pd.DataFrame:
        """Concatenates two DataFrames and adds a row for the 'Tipo' at the start of the DataFrame."""
        # Guardar orden original
        column_order = df1.iloc[:, 0].tolist()
        df1 = pd.concat([df1, pd.DataFrame([['Tipo'] + [df1_name] * (len(df1.columns)-1)], columns=df1.columns)]).reset_index(drop=True)
        df2 = pd.concat([df2, pd.DataFrame([['Tipo'] + [df2_name] * (len(df2.columns)-1)], columns=df2.columns)]).reset_index(drop=True)
        first_col_df1 = df1.columns[0]
        first_col_df2 = df2.columns[0]

        # Check if the first columns match before merging
        if first_col_df1 != first_col_df2:
            raise KeyError(f"The first columns do not match: '{first_col_df1}' and '{first_col_df2}'")
        result_df = pd.merge(df1, df2, on=first_col_df1, how='outer', suffixes=('_1', '_2'))

        # Convertir la fila Tipo en los nombres de las columnas
        tipo_row = result_df[result_df.iloc[ :,0] == 'Tipo'].iloc[0]
        nombre_original = result_df.columns[0]
        result_df.columns = tipo_row
        result_df.columns.values[0] = nombre_original

        # Preservar orden original
        result_df = result_df.set_index(result_df.columns[0]).reindex(column_order).reset_index()
        
        # Eliminar la fila "Tipo"
        result_df = result_df[result_df.iloc[:, 0] != 'Tipo'].reset_index(drop=True)

        return result_df
    
    def concat_multiple_dataframes(
        self,
        dfs: list[pd.DataFrame],
        df_names: list[str]
    ) -> pd.DataFrame:
        """
        Concatena múltiples DataFrames y añade sus nombres como identificadores.

        Parameters
        ----------
        dfs : list[pd.DataFrame]
            Lista de DataFrames a concatenar.
        df_names : list[str]
            Lista de nombres correspondientes a cada DataFrame.

        Returns
        -------
        pd.DataFrame
            DataFrame resultante con todos los datos combinados.

        Raises
        ------
        ValueError
            Si el número de DataFrames no coincide con el número de nombres o
            si hay menos de 2 DataFrames.
        KeyError
            Si los DataFrames no tienen la misma primera columna.
        """
        if len(dfs) != len(df_names):
            raise ValueError("El número de DataFrames debe coincidir con el número de nombres")
        
        if len(dfs) < 2:
            raise ValueError("Se necesitan al menos 2 DataFrames para concatenar")
        
        # Verificar que todos los DataFrames tengan la misma primera columna
        first_col = dfs[0].columns[0]
        for df in dfs[1:]:
            if df.columns[0] != first_col:
                raise KeyError(f"Todos los DataFrames deben tener el mismo nombre para la primera columna")
        
        # Renombrar las columnas para evitar duplicados
        for i, df in enumerate(dfs):
            df.columns = [f"{col}_{df_names[i]}" if col != first_col else col for col in df.columns]
        
        # Procesar cada DataFrame añadiendo la fila de tipo
        processed_dfs = []
        for df, name in zip(dfs, df_names):
            processed_df = pd.concat([
                df,
                pd.DataFrame([['Tipo'] + [name] * (len(df.columns)-1)], columns=df.columns)
            ]).reset_index(drop=True)
            processed_dfs.append(processed_df)

        # Guardar orden original
        column_order = processed_dfs[0].iloc[:, 0].tolist()

        # Realizar el merge de todos los DataFrames
        result_df = processed_dfs[0]
        for df in processed_dfs[1:]:
            result_df = pd.merge(
                result_df,
                df,
                on=first_col,
                how='outer',
                suffixes=('_left', '_right')
            )

        # Convertir la fila Tipo en los nombres de las columnas
        tipo_row = result_df[result_df.iloc[ :,0] == 'Tipo'].iloc[0]
        nombre_original = result_df.columns[0]
        result_df.columns = tipo_row
        result_df.columns.values[0] = nombre_original

        # Preservar orden original
        result_df = result_df.set_index(result_df.columns[0]).reindex(column_order).reset_index()

        # Eliminar la fila "Tipo"
        result_df = result_df[result_df.iloc[:, 0] != 'Tipo'].reset_index(drop=True)
        
        return result_df

    # Writing methods (simple)
    def dataframe_to_worksheet(self, df: pd.DataFrame, output_name: str, sheet_name: str = 'Hoja1', folder: str = "otros") -> None:
        """Writes a DataFrame to a worksheet in the Excel file.

        Parameters
        ----------
        df : pd.DataFrame
            The DataFrame to write to the worksheet.
        sheet_name : str, optional
            The name of the worksheet. Defaults to 'Hoja1'.
        folder : str, optional
            The name of the folder inside "products". Defaults to "otros".
        """
        output_file_path = os.path.join(self.output_path, folder, f'{output_name}.xlsx')
        with pd.ExcelWriter(output_file_path, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        
    def dataframes_to_worksheets(
        self,
        dfs: list[pd.DataFrame],
        output_name: str,
        sheet_names: list[str] = None, 
        skip_first: bool = True,
        folder: str = True
    ) -> None:
        """Writes multiple DataFrames to multiple worksheets in the Excel file.

        Parameters
        ----------
        dfs : List[pd.DataFrame]
            A list of DataFrames to write to the worksheets.
        sheet_names : list[str], optional
            A list of worksheet names. If not provided, default names will be used.
        skip_first : bool, optional
            Whether to start writing from Worksheet 2 onward. Defaults to True.
        folder : str, optional
            The name of the folder inside "products". Defaults to "otros".
        """
        if sheet_names is None:
            sheet_names = [f'Hoja{i+1}' for i in range(len(dfs))]  # Default sheet names: Hoja1, Hoja2, etc.

        if len(dfs) != len(sheet_names):
            raise ValueError("The number of DataFrames must match the number of sheet names.")

        output_file_path = os.path.join(self.output_path, folder, f'{output_name}.xlsx')
        os.makedirs(os.path.dirname(output_file_path), exist_ok= True)

        with pd.ExcelWriter(output_file_path, engine='openpyxl', mode='w') as writer:
            # If skip_first is True, create an empty sheet as the first one
            if skip_first:
                pd.DataFrame().to_excel(writer, sheet_name='Índice')

            # Write DataFrames to subsequent sheets
            for df, sheet_name in zip(dfs, sheet_names):
                df.to_excel(writer, sheet_name=sheet_name, index=False)
   
    # TODO: Use sheet index instead of name
