import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.styles.numbers import FORMAT_DATE_XLSX17
import datetime
from excel_automation.classes.formats.colors import Color
import os

script_dir = os.path.abspath(os.path.dirname(__file__))
#macros_folder = os.path.join(script_dir, "..", "macros", "excel")
save_dir = os.path.join(script_dir, "..", "charts")

class ExcelFormatter:
    def __init__(self, file_name: str = "", workbook: Workbook = None):
        """Class to open Excel files and apply beautiful format. 
        Engine: openpyxl

        Parameters
        ----------
        workbook : Workbook
            The workbook to be formatted.
        """
        self.file_path = os.path.join(script_dir, "..", "charts", f'{file_name}.xlsx')
        self.wb = openpyxl.load_workbook(self.file_path, data_only=True) if file_name else workbook
        # Access different worksheets with self.wb.sheetnames[int]
        self.ws = self.wb.active

    def _get_worksheet_position(self, position: str | int) -> int:
        if position == "start":
            index = 0 
        elif position == "end":
            index = len(self.wb.sheetnames)
        elif isinstance(position, int):
            if position < 0 or position > len(self.wb.sheetnames):
                raise ValueError(f"Invalid index {position}. Index must be between 0 and {len(self.wb.sheetnames)}.")
            index = position  # Add at the specified index
        else:
            raise ValueError(f"Invalid position '{position}'. Use 'start', 'end', or an integer index.")
        return index

    def get_worksheet_by_index(self, index: int):
        """Returns the worksheet at the specified index.
        Wrapper around self.wb.worsheets[index].

        Parameters
        ----------
        index : int
            The index of the worksheet to be accessed.

        Returns
        -------
        worksheet : Worksheet
            The worksheet at the specified index.
        """
        if 0 <= index < len(self.wb.worksheets):
            self.ws = self.wb.worksheets[index]
            return self.ws
        else:
            raise IndexError(f"Worksheet index {index} is out of range. The workbook has {len(self.wb.worksheets)} worksheets.")
        
    def create_new_worksheet(self, sheet_name: str = "Index", position: str | int = "end") -> None:
        """
        Creates a new worksheet in the workbook and adds it to the specified position.

        Parameters
        ----------
        sheet_name : str
            The name of the new worksheet to create. Defaults to "Index".
        position : str | int, optional
            The position where the new worksheet should be added. Options:
            - "start": Adds the worksheet as the first sheet.
            - "end": Adds the worksheet as the last sheet (default).
            - int: Adds the worksheet at the specified index (0-based).

        Returns
        -------
        None
        """
        if sheet_name in self.wb.sheetnames:
            raise ValueError(f"A worksheet with the name '{sheet_name}' already exists.")

        index = self._get_worksheet_position(position)

        # Create the new worksheet at the specified index
        self.wb.create_sheet(title=sheet_name, index=index)
        print(f"✅ Worksheet '{sheet_name}' created at position {index}.")
        

    def apply_database_format(self, position: str | int)-> None:
        """
        Applies a standardized database-like format to the current worksheet to improve visual appeal:

        - Sets column widths and hides gridlines.
        - Applies number formatting to numeric cells (with optional decimal places).
        - Formats the first column (assumed to contain years or dates) with a light gray background.
        - Formats the first row (assumed to contain headers) with a dark blue background, bold white text, and centered alignment.
        - Applies light border to data containing cells.

        Parameters
        ----------
        position : str | int, optional
            The position where the new worksheet should be added. Options:
            - "start": Adds the worksheet as the first sheet.
            - "end": Adds the worksheet as the last sheet (default).
            - int: Adds the worksheet at the specified index (0-based).

        Returns
        -------
        None
        """
        index = self._get_worksheet_position(position)
        self.get_worksheet_by_index(index) # Updates self.ws

        self.ws.column_dimensions['A'].width = 15
        self.ws.sheet_view.showGridLines = False
        w = Color.WHITE.no_hash
        g = Color.GRAY_LIGHT.no_hash
        b = Color.BLUE_DARK.no_hash

        if not isinstance(self.ws['B2'].value, (int, float)):
            print(f'La celda B2 de la hoja {index+1}  no contiene un valor de tipo número. Verifica el formato de las celdas')
        decimals = isinstance(self.ws['B2'].value, float)

        # Data iteration: number formatting to cells containing data
        vanilla_border = Border(left=Side(style='thin', color=g), right=Side(style='thin', color=g), top=Side(style='thin', color=g), bottom=Side(style='thin', color=g))
        for row in self.ws.iter_rows(min_row=2, max_row=self.ws.max_row, min_col=2, max_col=self.ws.max_column):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00" if decimals else "0"
                    cell.border = vanilla_border

        # First column (years)
        fill = PatternFill(start_color=g, fill_type="solid")
        custom_border = Border(left=Side(style='thin', color=w), right=Side(style='thin', color=w), top=Side(style='thin', color=w), bottom=Side(style='thin', color=w))
        for row in self.ws.iter_rows(min_row=1, max_row=self.ws.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.fill = fill
                cell.border = custom_border
                # Check if the cell contains a datetime value
                if isinstance(cell.value, (datetime.date, datetime.datetime)):
                    cell.number_format = FORMAT_DATE_XLSX17  # Apply mmm-yy format

        # First row (categories)
        fill = PatternFill(start_color=b, fill_type="solid")
        for row in self.ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=self.ws.max_column):
            for cell in row:
                cell.fill = fill
                cell.font = Font(color=w, bold=True)
                cell.border = custom_border
                cell.alignment = Alignment(horizontal= "center", vertical="center")  # Ajustar texto. TODO: no funciona en excel en línea


    def apply_database_format_all(self) -> None:
        """
        Applies the standardized database-like format to all worksheets in the workbook:

        - Sets column widths and hides gridlines.
        - Applies number formatting to numeric cells (with optional decimal places).
        - Formats the first column (assumed to contain years or dates) with a light gray background.
        - Formats the first row (assumed to contain headers) with a dark blue background, bold white text, and centered alignment.
        - Applies light border to data containing cells.
        
        Returns
        -------
        None
        """
        for index in range(len(self.wb.worksheets)): 
            self.apply_database_format(index)


    def save_workbook(self, output_name: str)-> None:
        """Save your workbook. Automatically includes extension in the name if not declared.

        Args:
            name (str, optional): Choose a name for your Excel file. Defaults to "excel_test".
        """
        
        output_path = os.path.join(save_dir, f'{output_name}.xlsx') if output_name else self.file_path
        self.wb.save(output_path)
        print(f'✅ Excel guardado como "{output_name}"')
