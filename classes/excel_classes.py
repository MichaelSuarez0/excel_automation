from win32com import client
import win32com.client as win32
import re
import time
import os
import pythoncom
from openpyxl import load_workbook, Workbook
import pandas as pd
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.axis import ChartLines
from openpyxl.drawing.fill import ColorChoice
from openpyxl.drawing.fill import SolidColorFillProperties
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.marker import Marker
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.chart.label import DataLabelList
from enum import Enum
from typing import Tuple, List, Optional
import xlwings as xw
from icecream import ic


class Color(Enum):
    RED: str = "d81326"
    RED_LIGHT: str = "FFABAB"
    BLUE_LIGHT: str = "A6CAEC"
    BLUE: str = "2E70D0"
    BLUE_DARK: str = "12213b"
    GRAY_LIGHT: str = "ebebeb"
    GRAY: str = "ebebe0"
    YELLOW: str = "FFC000"
    WHITE: str = 'FFFFFF'

script_dir = os.path.abspath(os.path.dirname(__file__))
macros_folder = os.path.join(script_dir, "..", "macros", "excel")
save_dir = os.path.join(script_dir, "..", "charts")

class ExcelOpenPyXL:
    def __init__(self, file_path):
        self.file_path = os.path.join(script_dir, "..", "databases", file_path)
        self.workbooks = {}  # Dictionary to store dynamically created workbooks (keys: names, values: Workbook objects)
        self.wb_count = 1  # Counter to track new workbooks
        self.wb = None
        self.ws = None
    
    def open_existing_workbook(self):
        self.wb = openpyxl.load_workbook(self.file_path)
        self.ws = self.wb.active
    
    def open_new_workbook(self) -> Tuple[Workbook, Worksheet]:
        """Dynamically create new workbooks and name them wb2, wb3, etc."""        
        self.wb_count += 1  
        
        # Create new workbook and assign it dynamically
        new_wb_name = f"wb{self.wb_count}"
        self.workbooks[new_wb_name] = Workbook()
        
        # Create new variables dynamically (starting with .self)
        setattr(self, new_wb_name, self.workbooks[new_wb_name])
        setattr(self, f"ws{self.wb_count}", self.workbooks[new_wb_name].active)

        print(f"✅ Created new workbook: {new_wb_name}")
        return self.workbooks[new_wb_name], self.workbooks[new_wb_name].active

    # TODO: No incluir extensión en el nombre del archivo
    def save_workbook(self, wb: Workbook = None, name: str = "excelpy.xlsx")-> None:
        # Save changes
        wb.save(os.path.join(save_dir, name)) if wb else self.wb.save(os.path.join(save_dir, name))
        print(f'✅ Excel guardado como "{name}"')
    
    @property
    def sheet_names(self) -> list: 
        """Devuelve una lista de los nombres de las hojas."""
        print("Sheet names:")
        for sheet_name in self.wb.sheetnames:
            print(f"- {sheet_name}")
        return self.wb.sheetnames

    @property
    def count_sheets(self) -> int:
        count = len(self.wb.sheetnames)
        print(f'The workbook has {count} sheets.')
        return count
    
    def excel_to_dataframe(self, sheet_name=None):
        if sheet_name is None:
            sheet_name = self.wb.sheetnames[0]  # Default to the first sheet if no sheet name is provided
        df = pd.read_excel(self.file_path, sheet_name=sheet_name)
        return df
    
    def dataframe_to_excel(self, df: pd.DataFrame, sheet_name='Hoja1', mode='w'):
        with pd.ExcelWriter(self.file_path, engine='openpyxl', mode=mode) as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False) # Pandas automatically saves
   
    # TODO: All that is missing is FUENTE and URL
    def apply_database_format(self, cws: Worksheet, sheet_name='Hoja1')-> None:
        #cws.column_dimensions['A'].width = 20
        cws.sheet_view.showGridLines = False
        w = Color.WHITE.value
        g = Color.GRAY_LIGHT.value

        # Data iteration: apply number formatting to all numeric cells in the worksheet
        vanilla_border = Border(left=Side(style='thin', color=g), right=Side(style='thin', color=g), top=Side(style='thin', color=g), bottom=Side(style='thin', color=g))
        for row in cws.iter_rows(min_row=2, max_row=cws.max_row, min_col=2, max_col=cws.max_column):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00"  # Max 2 decimals
                    cell.border = vanilla_border

        # First column
        fill = PatternFill(start_color=Color.GRAY_LIGHT.value, fill_type="solid")
        custom_border = Border(left=Side(style='thin', color=w), right=Side(style='thin', color=w), top=Side(style='thin', color=w), bottom=Side(style='thin', color=w))
        for row in cws.iter_rows(min_row=1, max_row=cws.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.fill = fill
                cell.border = custom_border

        # First row
        fill = PatternFill(start_color=Color.BLUE_DARK.value, fill_type="solid")
        for row in cws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=cws.max_column):
            for cell in row:
                cell.fill = fill
                cell.font = Font(color=w, bold=True)
                cell.border = custom_border
                cell.alignment = Alignment(horizontal= "center", vertical="center")  # Ajustar texto. TODO: no funciona en excel en línea

# TODO: Set axis max and min range dynamically
# TODO: Processes like iterating over selected rows can be modularized    
class ExcelAutoChart(ExcelOpenPyXL):
    def __init__(self, file_path):
        super().__init__(file_path)
        self.orientation = None  
    
    def copy_selected_cells(
        self,
        ws_title: str,
        selected_labels: List[str],
        target_workbook: Workbook = None,
    ) -> Tuple[Workbook, Worksheet]:
        """Method that handles orientation detection and data copying to a new worksheet (and workbook if specified)"""
        # 1. Orientation Detection Logic 
        is_vertical = False
        is_horizontal = False
        
        # Check column A for vertical categories
        if self.ws.cell(row=2, column=1).value and isinstance(self.ws.cell(row=2, column=1).value, str):
            is_vertical = True
            
        # Check row 1 for horizontal categories
        if self.ws.cell(row=1, column=2).value and isinstance(self.ws.cell(row=1, column=2).value, str):
            is_horizontal = True

        # Validate orientation
        if is_vertical and is_horizontal:
            raise ValueError("Categories in both rows and columns")
        elif not is_vertical and not is_horizontal:
            raise ValueError("No categories found")
        
        # Set orientation
        self.orientation = 'vertical' if is_vertical else 'horizontal'
        
        # 2. Category Collection Logic
        category_map = {}
        if self.orientation == 'vertical':
            for row in range(2, self.ws.max_row + 1):
                name = self.ws.cell(row, 1).value
                if name: category_map[name] = row
        else:
            for col in range(2, self.ws.max_column + 1):
                name = self.ws.cell(1, col).value
                if name: category_map[name] = col
        
        # 3. Get Selected Categories
        selected_categories = []
        if selected_labels:
            selected_categories = [category_map[label] for label in selected_labels if label in category_map]
        else:
            selected_categories = list(category_map.values())
        
        if not selected_categories:
            print("⚠️ No matching labels found")
            return

        # 4. Data Copying Logic
        if not target_workbook:
            new_wb, new_ws = self.open_new_workbook()
        else:
            new_wb = target_workbook
            new_ws = new_wb.create_sheet(title=ws_title)
            print(new_ws)
        

        if self.orientation == 'vertical':
            # Vertical data copy
            for col in range(1, self.ws.max_column + 1):
                new_ws.cell(1, col, self.ws.cell(1, col).value)
            for idx, row in enumerate(selected_categories, 2):
                for col in range(1, self.ws.max_column + 1):
                    new_ws.cell(idx, col, self.ws.cell(row, col).value)
        else:
            # Horizontal data copy
            for row in range(1, self.ws.max_row + 1):
                new_ws.cell(row, 1, self.ws.cell(row, 1).value)
            for idx, col in enumerate(selected_categories, 2):
                for row in range(1, self.ws.max_row + 1):
                    new_ws.cell(row, idx, self.ws.cell(row, col).value)

        return new_wb, new_ws, selected_categories

    
    def _prepare_chart_data(
        self,
        selected_labels: Optional[List[str]] = None,
        target_workbook: Workbook = None,
        ws_title: str = "ChartData"
    ) -> Tuple[Workbook, Worksheet, List, List, List[List]]:
        """
        Centralized method to handle all data preparation:
        1. Determines orientation
        2. Gets selected categories
        3. Copies relevant data to new worksheet
        4. Extracts dates, labels, and data series
        """
        # Copy to new worksheet and get selected categories
        new_wb, new_ws, selected_categories = self.copy_selected_cells(ws_title, selected_labels, target_workbook)
        
        # Extract dates and labels based on orientation
        if self.orientation == 'vertical':
            dates = [self.ws.cell(row=1, column=col).value 
                    for col in range(2, self.ws.max_column + 1)]
            labels = [self.ws.cell(row=idx, column=1).value 
                     for idx in selected_categories]
        else:
            dates = [self.ws.cell(row=row, column=1).value 
                    for row in range(2, self.ws.max_row + 1)]
            labels = [self.ws.cell(row=1, column=idx).value 
                     for idx in selected_categories]

        # Extract data series
        data_series = []
        for idx in selected_categories:
            if self.orientation == 'vertical':
                series = [self.ws.cell(row=idx, column=col).value 
                         for col in range(2, self.ws.max_column + 1)]
            else:
                series = [self.ws.cell(row=row, column=idx).value 
                         for row in range(2, self.ws.max_row + 1)]
            data_series.append(series)

        return new_wb, new_ws, dates, labels, data_series


    def create_line_chart(
        self,
        ws_title: str,
        source_ws: int = 0,
        selected_labels: list[str] | None = None,
        output_file: str = "line_chart.xlsx",
        marker: bool = True,
        target_workbook: Workbook = None,
    ) -> Tuple[Workbook, Worksheet]:
        """Creates a line chart using selected rows of data.

        Parameters
        ----------
        ws_title : str
            The name of the new worksheet to be created in the target workbook.
        
        source_ws : int, optional
            The worksheet number from which to extract data.

        selected_labels : list[str] | None, optional
            A list of row labels (first column values) to include in the chart.
            If None, all available categories are used, by default None.

        output_file : str, optional
            The name of the output Excel file containing the chart, by default "line_chart.xlsx".

        marker : bool, optional
            Whether to display markers on the line chart. If True, circular markers will be added to 
            each data point, by default True.

        target_workbook : Workbook, optional
            An existing workbook to which the chart will be added. If None, a new workbook is created.

        Returns
        -------
        Tuple[Workbook, Worksheet]
            The workbook and worksheet containing the chart.
        """
        self.open_existing_workbook()
        self.ws = self.wb.worksheets[source_ws]

        # Copy to new or selected workbook and get required data
        new_wb, new_ws, dates, labels, data_series = self._prepare_chart_data(selected_labels, target_workbook, ws_title)

        # Create the line chart
        chart = LineChart()
        chart.title = "Porcentaje de Cobertura por Departamento" 
        chart.style = 10  
        chart.width = 15  
        chart.height = 10 
        chart.legend.position = "b"  # Move legend to the bottom
        chart.graphical_properties = GraphicalProperties()
        chart.graphical_properties.ln = LineProperties(noFill=True)  # Removes the chart border

        # Add each selected category as a separate data series
        for idx, (label, series_data) in enumerate(zip(labels, data_series)):
            # Create a data series
            values = Reference(new_ws, min_col=idx + 2, min_row=2, max_row=len(dates) + 1)
            series = Series(values, title=label)

            # Define color and line style
            if idx == 0:
                color_code = Color.RED.value
                prstdash = "sysDash"
            elif idx == 1:
                color_code = Color.BLUE.value
                prstdash = "sysDot"
            elif idx == 2:
                color_code = Color.BLUE_DARK.value
                prstdash = "solid"

            # Define line properties
            series.graphicalProperties.ln = LineProperties(
                solidFill=ColorChoice(srgbClr=color_code),  # Alternating colors
                w=23000,  
                prstDash= prstdash
            ) 

            if marker == True:
                # Set marker (circle)
                series.marker = Marker(symbol="circle")
                series.marker.graphicalProperties = GraphicalProperties(
                    solidFill=ColorChoice(srgbClr=color_code),
                    noFill=False,
                    ln=LineProperties(
                        solidFill=ColorChoice(srgbClr=color_code)
                        ),
                )
    
            chart.append(series)

        # X-axis settings (years)
        categories = Reference(new_ws, min_col=1, max_col=1, min_row=2, max_row=new_ws.max_row) # Set X-axis categories from column 1
        chart.set_categories(categories) 
        chart.x_axis.delete = False  # ensure it's not hidden
        chart.x_axis.title = None  
        #chart.x_axis.reverseOrder = True # Descendent order
        chart.x_axis.minorTickMark = "out"  # ensure tick marks appear
        chart.x_axis.tickLblPos = "low"  # move labels to bottom
        
        # Y-axis settings
        chart.y_axis.delete = False  # ensure it's not hidden
        chart.y_axis.title = "Porcentaje (%)"
        chart.y_axis.scaling.max = 10 
        chart.y_axis.number_format = "0.0"  # one decimal
        chart.y_axis.tickLblPos = "low" # move labels closer to axis
        chart.y_axis.reverseOrder = True # Descendent order
        chart.y_axis.majorGridlines = ChartLines()
        chart.y_axis.majorGridlines.spPr = GraphicalProperties(ln=LineProperties(solidFill=Color.GRAY_LIGHT.value))  # Gray gridlines

        # Manual Layout for Chart Size & Position
        chart.layout = Layout(
            manualLayout=ManualLayout(
                x=0.00,  # Move right
                y=0.00,   # Move down
                h=0.85,   # Height
                w=0.9    # Width
            )
        )

        # TODO: Obtain position dynamically
        # Place chart
        new_ws.add_chart(chart, "P5") 

        # Apply format
        self.apply_database_format(new_ws)
        print(f"✅ Gráfico de líneas agregado")
        return new_wb, new_ws


    # TODO: Fix chart border not being deleted
    def create_vertical_bar_chart(
        self,
        ws_title: str,
        source_ws: int = 0,
        selected_labels: list[str] | None = None,
        output_file: str = "vertical_bar.xlsx",
        grouping: str = "standard",
        target_workbook: Workbook = None,
    ) -> Tuple[Workbook, Worksheet]:
        """Creates a vertical bar chart (stacked or grouped) using selected data rows.

        Parameters
        ----------
        ws_title : str
            The name of the new worksheet to be created in the target workbook.

        source_ws : int, optional
        The worksheet number from which to extract data.

        selected_labels : list[str] | None, optional
            A list of row labels (first column values) to include in the chart.
            If None, all available categories are used, by default None.

        output_file : str, optional
            The name of the output Excel file containing the chart, by default "vertical_bar.xlsm".

        grouping : str, optional
            The type of bar chart grouping (default standard). Options:
            - "standard" : Side-by-side bars for each category.
            - "stacked" : Stacked bars (sum of series per category).
            - "percentStacked" : 100% stacked bars (proportions per category).

        target_workbook : Workbook, optional
            An existing workbook to which the chart will be added. If None, a new workbook is created.

        Returns
        -------
        Tuple[Workbook, Worksheet]
            The workbook and worksheet containing the chart.
        """
        self.open_existing_workbook()
        self.ws = self.wb.worksheets[source_ws]

        # Copy to new or selected workbook and get required data
        new_wb, new_ws, dates, labels, data_series = self._prepare_chart_data(selected_labels, target_workbook, ws_title)

        # Chart creation and settings
        chart = BarChart()
        chart.title = "Porcentaje de Cobertura por Departamento" 
        chart.overlap = 100 
        chart.style = 10  
        chart.width = 15  
        chart.height = 10 
        chart.graphical_properties = GraphicalProperties()
        chart.graphical_properties.ln = LineProperties(noFill=True)  # Removes the chart border
        if not grouping:
            chart.grouping = "standard" # Stacked bar chart
        else:
            chart.grouping = grouping

        # Add each selected category as a separate data series
        for idx, (label, series_data) in enumerate(zip(labels, data_series)):
            # Create a data series
            values = Reference(new_ws, min_col=idx + 2, min_row=2, max_row=len(dates) + 1)
            series = Series(values, title=label)
            
            # Alternate bar colors
            color_code = Color.BLUE_DARK.value if idx % 2 == 0 else Color.BLUE.value 

            # Aplicar color de relleno a la serie (solo para gráficos de barras)
            series.graphicalProperties.solidFill = ColorChoice(srgbClr=color_code)
            chart.append(series)

        # X-axis settings
        categories = Reference(new_ws, min_col=2, max_col=new_ws.max_column, min_row=1, max_row=1)
        chart.set_categories(categories) # Set X-axis categories from Row 1
        chart.x_axis.delete = False  # ensure it's not hidden
        chart.x_axis.title = None  # Remove X-axis title
        chart.x_axis.minorTickMark = "out"  # ensure tick marks appear
        chart.x_axis.tickLblPos = "low"  # move labels to bottom

        # Y-axis settings
        chart.y_axis.delete = False  # ensure it's not hidden
        chart.y_axis.title = "Porcentaje (%)"
        chart.y_axis.scaling.max = 100 
        chart.y_axis.scaling.min = 0
        chart.y_axis.number_format = "0"  # remove decimals
        chart.y_axis.tickLblPos = "low" # move labels closer to axis

        # Move legend to avoid overlap
        chart.legend.position = "b"  # Move legend to the bottom

        # Apply gray gridlines
        chart.y_axis.majorGridlines = ChartLines()
        chart.y_axis.majorGridlines.spPr = GraphicalProperties(ln=LineProperties(solidFill=Color.GRAY_LIGHT.value))  # Gray gridlines
        # chart.x_axis.majorGridlines = ChartLines()
        # chart.x_axis.majorGridlines.spPr = GraphicalProperties(ln=LineProperties(solidFill=Color.LIGHT_GRAY.value))  # Gray gridlines

        # Manual Layout for Chart Size & Position
        chart.layout = Layout(
            manualLayout=ManualLayout(
                x=0.00,  # Move right
                y=0.01,   # Move down
                h=0.78,   # Height
                w=0.9    # Width
            )
        )

        # Place chart
        new_ws.add_chart(chart, "P5") 

        # Apply format
        self.apply_database_format(new_ws)

        print(f"✅ Gráfico agregado a '{output_file}' con los datos seleccionados.")
        return new_wb, new_ws


    # TODO: highlighted_labels approach does not work, must be done with macros
    def create_horizontal_bar_chart(
        self,
        ws_title: str,
        source_ws: int = 0,
        highlighted_labels: list[str] | None = None,
        output_file: str = "horizontal_bar.xlsx",
        target_workbook: Workbook = None
    ) -> Tuple[Workbook, Worksheet]:
        """Creates a line chart using selected rows of data.

        Parameters
        ----------
        ws_title : str
            The name of the new worksheet to be created in the target workbook.

        source_ws : int
        The worksheet number from which to extract data.

        highlighted_labels : list[str] | None, optional
            A list of row labels (first column values) to highlight in the chart.
            If None, no labels are highlighted.

        output_file : str, optional
            The name of the output Excel file containing the chart, by default "horizontal_bar.xlsx".

        target_workbook : Workbook, optional
            An existing workbook to which the chart will be added. If None, a new workbook is created.

        Returns
        -------
        Tuple[Workbook, Worksheet]
            The workbook and worksheet containing the chart.
        """
        self.open_existing_workbook()
        self.ws = self.wb.worksheets[source_ws]

        # Copy to new or selected workbook and get required data
        new_wb, new_ws, dates, labels, data_series = self._prepare_chart_data(target_workbook=target_workbook, ws_title= ws_title)
        
        # Chart creation and settings
        chart = BarChart()
        chart.type = "bar"  # Horizontal bar chart
        chart.gapWidth = 20  # Ajusta el espacio entre barras
        chart.title = "Porcentaje de Cobertura por Departamento" 
        chart.style = 10  
        chart.width = 15  
        chart.height = 15 
        chart.legend = None  # Remove legend
        chart.graphical_properties = GraphicalProperties()
        chart.graphical_properties.ln = LineProperties(noFill=True)  # Removes the chart border

        # # Define data range (column 2 onward)
        data = Reference(new_ws, min_col=2, max_col=new_ws.max_column, min_row=2, max_row=new_ws.max_row)
        series = Series(data, title="Valores")  # No title to avoid legend entries
        series.graphicalProperties.solidFill = ColorChoice(srgbClr=Color.BLUE_DARK.value)  # Blue color
        # Data labels
        series.dLbls = DataLabelList(showVal=True, dLblPos= "outEnd", showCatName = False, showSerName = False, showLegendKey = False)  # Mostrar valores en las etiquetas de datos

        # Append the series to the chart
        chart.append(series)

        # Define categories (column 1)
        categories = Reference(new_ws, min_col=1, min_row=2, max_row=new_ws.max_row)
        chart.set_categories(categories)

        # # Create separate series for highlighted and non-highlighted rows
        # for row_idx in range(2, new_ws.max_row + 1):  # Rows start at 2 (header is row 1)
        #     # Define data range for the current row
        #     data = Reference(new_ws, min_col=2, max_col=last_col, min_row=row_idx, max_row=row_idx)
            
        #     # Create a series for the current row (without a title)
        #     series = Series(data, title=None)  # No title to avoid legend entries
            
        #     # Set color based on whether the row is highlighted
        #     if row_idx in highlighted_rows:
        #         series.graphicalProperties.solidFill = ColorChoice(srgbClr=Color.RED.value)  # Red color
        #     else:
        #         series.graphicalProperties.solidFill = ColorChoice(srgbClr=Color.DARK_BLUE.value)  # Blue color
            
        #     # Append the series to the chart
        #     chart.append(series)

        # Y-axis settings
        chart.x_axis.delete = False  # ensure it's not hidden
        chart.x_axis.title = None  # Remove X-axis title
        chart.x_axis.reverseOrder = True # Descendent order
        chart.x_axis.minorTickMark = "out"  # ensure tick marks appear
        chart.x_axis.tickLblPos = "low"  # move labels to bottom
        # chart.x_axis.majorGridlines = ChartLines()
        # chart.y_axis.majorGridlines.spPr = GraphicalProperties(ln=LineProperties(solidFill=Color.LIGHT_GRAY.value))  # Gray gridlines
        
        # X-axis settings
        chart.y_axis.delete = False  # ensure it's not hidden
        chart.y_axis.title = "Porcentaje (%)"
        chart.y_axis.scaling.max = 10 
        chart.y_axis.number_format = "0.0"  # one decimal
        chart.y_axis.tickLblPos = "low" # move labels closer to axis
        chart.y_axis.reverseOrder = True # Descendent order
        chart.y_axis.majorGridlines = ChartLines()
        chart.y_axis.majorGridlines.spPr = GraphicalProperties(ln=LineProperties(solidFill=Color.GRAY_LIGHT.value))  # Gray gridlines

        # Manual Layout for Chart Size & Position
        chart.layout = Layout(
            manualLayout=ManualLayout(
                x=0.00,  # Move right
                y=0.00,   # Move down
                h=0.9,   # Height
                w=0.9    # Width
            )
        )

        # TODO: Obtain position dynamically
        # Place chart
        new_ws.add_chart(chart, "P5") 

        # Apply format
        self.apply_database_format(new_ws)

        print(f"✅ Gráfico agregado a '{output_file}' con los datos seleccionados.")
        return new_wb, new_ws


class ExcelFormatter:
    def __init__(self, file_name):
        """
        Initializes the ExcelFormatter with the path to the Excel file and the macros folder.

        Parameters:
        -----------
        file_name : str
            The name of the Excel workbook.
        save_dir : str
            The directory where the workbook is saved.
        macros_folder : str
            The path to the folder containing the macro files.
        """
        self.file_path = os.path.join(save_dir, file_name)
        self.macros_folder = macros_folder
        self.app = None
        self.workbook = None

    # TODO: REVISAR    
    def convert_xlsx_to_xlsm(self, input_file, output_file):
        """
        Converts an .xlsx file to .xlsm format using xlwings.

        Parameters:
        -----------
        input_file : str
            The name of the input .xlsx file.
        output_file : str
            The name of the output .xlsm file.
        """
        try:
            # Open the .xlsx file
            self.app = xw.App(visible=False)  # Run Excel in the background
            self.workbook = self.app.books.open(os.path.join(self.save_dir, input_file))
            # Save as .xlsm
            self.workbook.save(os.path.join(self.save_dir, output_file))
            print(f"File saved as '{output_file}'.")
        except Exception as e:
            print(f"Error converting file: {e}")
        finally:
            # Close the workbook and quit the app
            if self.workbook:
                self.workbook.close()
            if self.app:
                self.app.quit()

    def open_workbook(self):
        """
        Opens the Excel workbook and initializes the Excel application.
        """
        self.app = xw.App(visible=False)  # Run Excel in the background
        self.workbook = self.app.books.open(self.file_path)

    def close_workbook(self):
        """
        Saves and closes the workbook, and quits the Excel application.
        """
        if self.workbook:
            self.workbook.save()
            self.workbook.close()
        if self.app:
            self.app.quit()

    def load_macro_from_file(self, macro_name):
        """
        Loads a VBA macro from a file.

        Parameters:
        -----------
        macro_name : str
            The name of the macro (and the file name without extension).

        Returns:
        --------
        str
            The VBA code for the macro.
        """
        macro_file = os.path.join(self.macros_folder, f"{macro_name}.bas")
        try:
            with open(macro_file, "r") as file:
                macro_code = file.read()
            return macro_code
        except Exception as e:
            print(f"Error loading macro '{macro_name}' from file: {e}")
            return None

    def add_macro(self, macro_name):
        """
        Adds a VBA macro to the workbook by loading it from a file.

        Parameters:
        -----------
        macro_name : str
            The name of the macro (and the file name without extension).
        """
        macro_code = self.load_macro_from_file(macro_name)
        if not macro_code:
            return

        try:
            # Add the macro to the workbook
            self.workbook.api.VBProject.VBComponents.Add(1).CodeModule.AddFromString(macro_code)
            print(f"Macro '{macro_name}' added successfully.")
        except Exception as e:
            print(f"Error adding macro: {e}")
            print("Please ensure 'Trust access to the VBA project object model' is enabled in Excel.")

    def run_macro(self, macro_name):
        """
        Runs a VBA macro in the workbook.

        Parameters:
        -----------
        macro_name : str
            The name of the macro to run.
        """
        try:
            # Run the macro
            self.workbook.macro(macro_name)()
            print(f"Macro '{macro_name}' executed successfully.")
        except Exception as e:
            print(f"Error running macro: {e}")

    def remove_chart_shadows(self):
        """
        Adds and runs a macro to remove shadows from all charts in the workbook.
        The macro is loaded from a file named 'RemoveChartShadows.bas'.
        """
        macro_name = "RemoveChartShadows"

        try:
            self.open_workbook()
            self.add_macro(macro_name)
            self.run_macro(macro_name)
        finally:
            # Close the workbook
            self.close_workbook()