import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.styles.numbers import FORMAT_DATE_XLSX17
import datetime
from microsoft_office_automation.classes.colors import Color

class ExcelFormatter:
    def __init__(self, workbook: Workbook | None, file_name: str = ""):
        """Class to open Excel files and apply beautiful format. 
        Engine: openpyxl

        Parameters
        ----------
        workbook : Workbook
            The workbook to be formatted.
        """
        self.file_name = file_name
        self.wb = openpyxl.load_workbook(self.file_name) if file_name else workbook
        # Access different worksheets with self.wb.sheetnames[int]
        self.ws= self.wb.active
    
        
    def apply_database_format(self, sheet_name='Hoja1', decimals = True)-> None:
        self.ws.column_dimensions['A'].width = 15
        self.ws.sheet_view.showGridLines = False
        w = Color.WHITE.value
        g = Color.GRAY_LIGHT.value

        # Data iteration: number formatting to cells containing data
        vanilla_border = Border(left=Side(style='thin', color=g), right=Side(style='thin', color=g), top=Side(style='thin', color=g), bottom=Side(style='thin', color=g))
        for row in self.ws.iter_rows(min_row=2, max_row=self.ws.max_row, min_col=2, max_col=self.ws.max_column):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00" if decimals else "0"
                    cell.border = vanilla_border

        # First column (years)
        fill = PatternFill(start_color=Color.GRAY_LIGHT.value, fill_type="solid")
        custom_border = Border(left=Side(style='thin', color=w), right=Side(style='thin', color=w), top=Side(style='thin', color=w), bottom=Side(style='thin', color=w))
        for row in self.ws.iter_rows(min_row=1, max_row=self.ws.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.fill = fill
                cell.border = custom_border
                # Check if the cell contains a datetime value
                if isinstance(cell.value, (datetime.date, datetime.datetime)):
                    cell.number_format = FORMAT_DATE_XLSX17  # Apply mmm-yy format

        # First row (categories)
        fill = PatternFill(start_color=Color.BLUE_DARK.value, fill_type="solid")
        for row in self.ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=self.ws.max_column):
            for cell in row:
                cell.fill = fill
                cell.font = Font(color=w, bold=True)
                cell.border = custom_border
                cell.alignment = Alignment(horizontal= "center", vertical="center")  # Ajustar texto. TODO: no funciona en excel en línea

