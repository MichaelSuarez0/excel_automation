import re
import time
import os
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.styles.numbers import FORMAT_DATE_XLSX17
from enum import Enum
from icecream import ic
import logging
import datetime
import pandas as pd
from typing import Tuple, Optional
import xlsxwriter
from xlsxwriter.workbook import Workbook
from xlsxwriter.worksheet import Worksheet
from xlsxwriter.format import Format
from xlsxwriter.utility import xl_range

# Hex colors
class Color(Enum):
    RED: str = "#C80000"
    RED_LIGHT: str = "#FFABAB"
    BLUE_LIGHT: str = "#A6CAEC"
    BLUE: str = "#3B79D5"
    BLUE_DARK: str = "#12213b"
    GREEN_DARK: str = "#008E2C"
    GRAY_LIGHT: str = "#ebebeb"
    GRAY: str = "#ebebe0"
    YELLOW: str = "#FFC000"
    WHITE: str = '#FFFFFF'
    ORANGE: str = "#DD6909"


script_dir = os.path.abspath(os.path.dirname(__file__))
macros_folder = os.path.join(script_dir, "..", "macros", "excel")
save_dir = os.path.join(script_dir, "..", "charts")

class ExcelDataExtractor():
    def __init__(self, file_name: str, output_name: str):
        """Class to obtain data from an Excel file, convert to DataFrame, apply transformations, and export it. 
        Engine: mostly pandas

        Parameters
        ----------
        file_name : str
                The name of the Excel file to be loaded (from databases folder)
        """
        self.file_path = os.path.join(script_dir, "..", "databases", f'{file_name}.xlsx')
        self.output_path = os.path.join(script_dir, "..", "charts", f'{output_name}.xlsx')
        self.wb = None
        self.ws = None
        self.load_workbook()
    
    def load_workbook(self):
        self.wb = openpyxl.load_workbook(self.file_path)
        # Access different worksheets with self.wb.sheetnames[int]
        self.ws= self.wb.active
    
    def save_workbook(self)-> None:
        """Save your workbook. Automatically includes extension in the name if not declared.

        Args:
            name (str, optional): Choose a name for your Excel file. Defaults to "excel_test".
        """
        self.wb.save(self.output_path)
        print(f'✅ Excel guardado como "{self.output_path}"')

    # def open_new_workbook(self, ws_name: str = None) -> Tuple[Workbook, Worksheet]:
    #     """Dynamically create new workbooks and name them wb2, wb3, etc."""        
    #     self.wb_count += 1  
        
    #     # Create new workbook and assign it dynamically
    #     new_wb_name = f"wb{self.wb_count}"
    #     self.workbooks[(self.wb_count)] = Workbook()
        
    #     # Create new variables dynamically (starting with .self)
    #     setattr(self, new_wb_name, self.workbooks[self.wb_count])
    #     setattr(self, f"ws{self.wb_count}", self.workbooks[self.wb_count].active)
    #     # Get the active worksheet or create a new one with the specified name
    #     if ws_name:
    #         new_ws = self.workbooks[self.wb_count].create_sheet(title=ws_name)
    #     else:
    #         new_ws = self.workbooks[self.wb_count].active

    #     print(f"✅ Created new workbook: {new_wb_name}")
    #     return self.workbooks[self.wb_count], new_ws
    
    @property
    def sheet_names(self) -> list: 
        """Devuelve una lista de los nombres de las hojas."""
        print("Sheet names:")
        for sheet_name in self.wb.sheetnames:
            print(f"- {sheet_name}")
        return self.wb.sheetnames

    @property
    def count_sheets(self) -> int:
        count = len(self.wb.sheetnames)
        print(f'The workbook has {count} sheets.')
        return count
    
    # Opening methods
    def worksheet_to_dataframe(self, sheet_index: int = None) -> pd.DataFrame:
        """Reads sheet and return a DataFrame, may specify worksheet index"""
        sheet_name = self.wb.sheetnames[sheet_index] if sheet_index else self.wb.sheetnames[0]
        df = pd.read_excel(self.file_path, sheet_name)
        return df
    
    def worksheets_to_dataframes(self, include_first = False) -> list[pd.DataFrame]:
        """Reads all sheets at once and returns a list of DataFrames, may specify to skip first"""
        dfs_dict = pd.read_excel(self.file_path, sheet_name=None) # This method reads all sheets at once a returns a dictionary of DataFrames
        sheet_names = list(dfs_dict.keys())[1:] if not include_first else list(dfs_dict.keys())
        dfs = [dfs_dict[name] for name in sheet_names]
        return dfs
    
    # Transformation methods
    def normalize_orientation(self, dfs: pd.DataFrame | list[pd.DataFrame]) -> list[pd.DataFrame]:
        """Normalizes the orientation of all DataFrames. Converts to list if a single df is provided"""
        if not isinstance(dfs, (pd.DataFrame, list)):
            raise ValueError("Must provide either a DataFrame or a list of DataFrames")
        if isinstance(dfs, pd.DataFrame):
            dfs= [dfs]
        normalized_dfs = []
        for df in dfs:
            # Check if the first row contains categories. If it does, it will transpose the df.
            if isinstance(df.iloc[0, 1], str) and isinstance(df.iloc[1, 0], str):
                continue
            if not isinstance(df.iloc[0, 1], str):
                index_name = df.columns[0]
                df = df.set_index(df.columns[0]).transpose() # Manually set another index, or else the default index stays on top
                df.reset_index(inplace=True)
                df.columns = [index_name] + df.columns[1:].tolist()  # I loathe pandas indexes
            normalized_dfs.append(df)
        
        return normalized_dfs
    
    def filter_data(
        self,
        df: pd.DataFrame,
        selected_categories: Optional[list[str]] = None,
    ) -> pd.DataFrame:
        """Filters data based on selected_categories"""
        if selected_categories:
            cols = [df.columns[0]] # Start with the first column, remember labels are in Row 1
            
            # Loop through the remaining columns and add them if they are in selected_labels
            for col in df.columns[1:]:
                if col in selected_categories:
                    cols.append(col)
            filtered_df = df[cols]
        else:
            filtered_df = df
        #ic(filtered_df)

        return filtered_df
    
    # Writing methods (simple)
    def dataframe_to_worksheet(self, df: pd.DataFrame, sheet_name: str = 'Hoja1', mode: str = 'w') -> None:
        """Writes a DataFrame to a worksheet in the Excel file.

        Parameters
        ----------
        df : pd.DataFrame
            The DataFrame to write to the worksheet.
        sheet_name : str, optional
            The name of the worksheet. Defaults to 'Hoja1'.
        mode : str, optional
            The mode to open the Excel file ('w' for write, 'a' for append). Defaults to 'w'.
        """
        with pd.ExcelWriter(self.output_path, engine='openpyxl', mode=mode) as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        
    def dataframes_to_worksheets(self, dfs: list[pd.DataFrame], sheet_names: list[str] = None, mode: str = 'w', skip_first: bool = True) -> None:
        """Writes multiple DataFrames to multiple worksheets in the Excel file.

        Parameters
        ----------
        dfs : List[pd.DataFrame]
            A list of DataFrames to write to the worksheets.
        sheet_names : list[str], optional
            A list of worksheet names. If not provided, default names will be used.
        mode : str, optional
            The mode to open the Excel file ('w' for write, 'a' for append). Defaults to 'w'.
        skip_first : bool, optional
            Whether to start writing from Worksheet 2 onward. Defaults to True.
        """
        if sheet_names is None:
            sheet_names = [f'Hoja{i+1}' for i in range(len(dfs))]  # Default sheet names: Hoja1, Hoja2, etc.

        if len(dfs) != len(sheet_names):
            raise ValueError("The number of DataFrames must match the number of sheet names.")

        # If skip_first is True, add a blank worksheet as the first one
        if skip_first:
            with pd.ExcelWriter(self.output_path, engine='openpyxl', mode=mode) as writer:
                pd.DataFrame().to_excel(writer, sheet_name='Índice') 

        # Write DataFrames to subsequent sheets
        for i, (df, sheet_name) in enumerate(zip(dfs, sheet_names), start=1 if skip_first else 0):
            self.dataframe_to_worksheet(df, sheet_name=sheet_name, mode=mode)
   
    # TODO: All that is missing is FUENTE and URL
    # TODO: Use sheet index instead of name

class ExcelFormatter:
    def __init__(self, workbook: Workbook | None, file_name: str = ""):
        """Class to open Excel files and apply beautiful format. 
        Engine: openpyxl

        Parameters
        ----------
        workbook : Workbook
            The workbook to be formatted.
        """
        self.file_name = file_name
        self.wb = openpyxl.load_workbook(self.file_name) if file_name else workbook
        # Access different worksheets with self.wb.sheetnames[int]
        self.ws= self.wb.active
    
        
    def apply_database_format(self, sheet_name='Hoja1', decimals = True)-> None:
        self.ws.column_dimensions['A'].width = 15
        self.ws.sheet_view.showGridLines = False
        w = Color.WHITE.value
        g = Color.GRAY_LIGHT.value

        # Data iteration: number formatting to cells containing data
        vanilla_border = Border(left=Side(style='thin', color=g), right=Side(style='thin', color=g), top=Side(style='thin', color=g), bottom=Side(style='thin', color=g))
        for row in self.ws.iter_rows(min_row=2, max_row=self.ws.max_row, min_col=2, max_col=self.ws.max_column):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00" if decimals else "0"
                    cell.border = vanilla_border

        # First column (years)
        fill = PatternFill(start_color=Color.GRAY_LIGHT.value, fill_type="solid")
        custom_border = Border(left=Side(style='thin', color=w), right=Side(style='thin', color=w), top=Side(style='thin', color=w), bottom=Side(style='thin', color=w))
        for row in self.ws.iter_rows(min_row=1, max_row=self.ws.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.fill = fill
                cell.border = custom_border
                # Check if the cell contains a datetime value
                if isinstance(cell.value, (datetime.date, datetime.datetime)):
                    cell.number_format = FORMAT_DATE_XLSX17  # Apply mmm-yy format

        # First row (categories)
        fill = PatternFill(start_color=Color.BLUE_DARK.value, fill_type="solid")
        for row in self.ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=self.ws.max_column):
            for cell in row:
                cell.fill = fill
                cell.font = Font(color=w, bold=True)
                cell.border = custom_border
                cell.alignment = Alignment(horizontal= "center", vertical="center")  # Ajustar texto. TODO: no funciona en excel en línea


# TODO: Set axis max and min range dynamically
# TODO: Parámetro para especificar si leyenda o no, si no, que se haga más largo el cuadro del gráfico (porque ya no habría leyenda)
# TODO: Reducir el chart area, chart layout
class ExcelAutoChart:
    def __init__(self, df_list: list[pd.DataFrame], output_name: str = "ExcelAutoChart"):
        """Class to write to Excel files from DataFrames and creating charts. Engine: xlsxwriter

        Parameters
        ----------
            df_list (list[pd.DataFrame]): Data that will be written to Excel
            output_name (str, optional): File name for the output file. Defaults to "ExcelAutoChart".
        """
        self.writer = pd.ExcelWriter(os.path.join(save_dir, f'{output_name}.xlsx'), engine='xlsxwriter')
        self.workbook: Workbook = self.writer.book
        self.df_list = df_list
        self._initialize_chart_formats()

    # TODO: Merge with _create_base_chart
    def _initialize_chart_formats(self):
        """Predefine chart-specific formats using the Color enum"""
        color_list = [Color.BLUE_DARK.value, Color.RED.value, Color.GREEN_DARK.value, Color.ORANGE.value, Color.GRAY.value]
        self.chart_formats = {
            'line': {
                'colors': color_list,
                'width': 2.5,
                'dash_types': ['solid', 'dash', 'dot', 'solid']
            },
            'marker': {
                'size': 6,
                'colors': color_list
            }
        }

    def _write_to_excel(self, df: pd.DataFrame, sheet_name: str = "ChartData") -> Tuple[pd.DataFrame, Worksheet]:
        """Process source data and prepare worksheet for charting"""
        df.to_excel(self.writer, sheet_name=sheet_name, index=False)
        worksheet = self.writer.sheets[sheet_name]
        return df, worksheet
    
    def _create_base_chart(self, worksheet: Worksheet, chart_type: str):
        """Default settings for all chart types"""
        chart = self.workbook.add_chart({'type': chart_type})

        chart.set_size({'width': 600, 'height': 420})
        chart.set_legend({'position': 'bottom'})
        chart.set_plotarea({
            'layout': {
                'x':      0.11,
                'y':      0.10,
                'width':  0.83,
                'height': 0.75,
            }
        })
        chart.set_chartarea({'border': {'none': True}})

         
    def create_line_chart(
        self,
        index: int = 0,
        sheet_name: str = "LineChart",
        marker: bool = True
    ) -> Worksheet:
        """Generate line chart with color scheme from Color enum"""
        color_list = [Color.BLUE_DARK.value, Color.RED.value, Color.GREEN_DARK.value, Color.ORANGE.value, Color.GRAY.value]
        data_df, worksheet = self._write_to_excel(self.df_list[index], sheet_name)
        
        # Check if the DataFrame is empty
        if data_df.empty:
            raise ValueError("DataFrame is empty. No data to plot.")
        
        chart = self._create_base_chart(worksheet, 'line')

        # Uncomment for IntelliSense
        #chart = self.workbook.add_chart({'type': 'line'})
        
        # Configure chart appearance
        # chart.set_title({'name': 'Performance Over Time', 'name_font': {'size': 14}})
        # chart.set_size({'width': 600, 'height': 420})
        # chart.set_legend({'position': 'bottom'})

        # Add data series with color scheme
        for idx, col in enumerate(data_df.columns[1:]):
            col_letter = chr(66 + idx)  # Get column letter (e.g., B, C, D, ...)
            #print(f"Adding series for column {col}: {col_letter}")  # Debug

            series_params = {
                'name': f"={sheet_name}!${col_letter}$1",  # Use column letter dynamically
                'categories': f"={sheet_name}!$A$2:$A${len(data_df)+1}",
                'values': f"={sheet_name}!${col_letter}$2:${col_letter}${len(data_df)+1}",
                'line': {
                    'color': self.chart_formats['line']['colors'][idx % len(self.chart_formats['line']['colors'])],
                    'width': self.chart_formats['line']['width'],
                    'dash_type': self.chart_formats['line']['dash_types'][idx % len(self.chart_formats['line']['dash_types'])],
                }
            }

            if marker:
                marker_color = self.chart_formats['marker']['colors'][idx % len(self.chart_formats['marker']['colors'])]
                series_params['marker'] = {
                    'type': 'circle',
                    'size': self.chart_formats['marker']['size'],
                    'fill': {'color': marker_color},
                    'line': {'color': marker_color},
                }

            chart.add_series(series_params)


        # Axis configuration
        chart.set_y_axis({
            'name': 'Percentage (%)',
            'num_format': '0.00',
            'major_gridlines': {
                'visible': True,
                'line': {'color': Color.GRAY_LIGHT.value}
            }
        })
        
        chart.set_x_axis({
            'name': '',
            'num_format': '0',
            'text_axis': True,
        })

        # Insert chart with proper positioning
        worksheet.insert_chart('E2', chart, {'x_offset': 25, 'y_offset': 10})
        
        self.writer.close()  # Automatically saves
        return worksheet
    
    def create_bar_chart(
        self,
        index: int = 0,
        sheet_name: str = "BarChart",
        grouping: str = "standard"
    ) -> Worksheet:
        """Generate vertical bar chart with color scheme from Color enum"""
        color_list = [Color.BLUE_DARK.value, Color.RED.value, Color.GREEN_DARK.value, Color.ORANGE.value, Color.GRAY.value]
        data_df, worksheet = self._write_to_excel(self.df_list[index], sheet_name)
        
        # Check if DataFrame is empty
        if data_df.empty:
            raise ValueError("DataFrame is empty. No data to plot.")

        # Map grouping types to xlsxwriter subtypes
        subtype_map = {
            'standard': 'clustered',
            'stacked': 'stacked',
            'percentStacked': 'percent_stacked'
        }
        subtype = subtype_map.get(grouping, 'clustered')

        # Create column chart (vertical bars)
        chart = self.workbook.add_chart({'type': 'column', 'subtype': subtype})
        
        # Configure chart appearance
        #chart.set_title({'name': 'Coverage Percentage by Department', 'name_font': {'size': 14}})
        chart.set_size({'width': 600, 'height': 380})
        chart.set_legend({'position': 'bottom'})

        # Get colors from predefined formats
        colors = self.chart_formats['line']['colors']

        # Add data series with color scheme
        for idx, col in enumerate(data_df.columns[1:]):
            col_idx = idx + 1  # Skip first column (categories)
            color = colors[idx % len(colors)] # are cycled through the predefined list of colors (colors), even if there are more series than colors.

            series_params = {
                'name': [sheet_name, 0, col_idx],  # Header row
                'categories': [sheet_name, 1, 0, len(data_df), 0],  # First column data
                'values': [sheet_name, 1, col_idx, len(data_df), col_idx],  # Value columns
                'fill': {'color': color},
            }

            chart.add_series(series_params)

        # Y-axis configuration
        chart.set_y_axis({
            'name': 'Percentage (%)',
            'num_format': '0',  # No decimals
            'max': 100,
            'min': 0,
            'major_gridlines': {
                'visible': True,
                'line': {'color': Color.GRAY_LIGHT.value}
            }
        })

        # X-axis configuration
        chart.set_x_axis({
            'name': '',
            'text_axis': True,  # Treat as text categories
            'num_format': '@',  # Text format
        })

        # chart.set_plotarea({
        #     'layout': {
        #         'x':      0.11,
        #         'y':      0.10,
        #         'width':  0.83,
        #         'height': 0.75,
        #     }
        # })

        # chart.set_chartarea({'border': {'none': True}})

        # Insert chart with proper positioning
        worksheet.insert_chart('E2', chart, {'x_offset': 25, 'y_offset': 10})
        
        self.writer.close()  # Save and close workbook
        return worksheet


# class ExcelAutomation:
#     def __init__(self, file_name: str):
#         """Class for automating Excel-related tasks.

#         Parameters
#         ----------
#         file_name : str
#             The name of the Excel file to be created or loaded.
#         """
#         self.handler = ExcelHandler(file_name)  # Initialize ExcelHandler
#         self.formatter = ExcelFormatter(workbook= self.handler.wb)  # Pass the workbook to ExcelFormatter

#     def save_workbook(self, name: str = "excel_test") -> None:
#         """Saves the workbook using ExcelHandler."""
#         self.handler.save_workbook(name)

#     def apply_database_format(self, sheet_name: str = 'Hoja1', decimals: bool = True) -> None:
#         """Applies database formatting using ExcelFormatter."""
#         self.formatter.apply_database_format(sheet_name, decimals)

#     def get_sheet_names(self) -> list[str]:
#         """Returns the sheet names using ExcelHandler."""
#         return self.handler.sheet_names

#     def get_count_sheets(self) -> int:
#         """Returns the number of sheets using ExcelHandler."""
#         return self.handler.count_sheets
